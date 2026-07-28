"""Read Amazon Ads search term reports from CSV or XLSX.

Amazon's report headers vary by marketplace and report window, e.g.
"7 Day Total Sales", "14 Day Total Sales (₹)", "Total Advertising Cost of
Sales (ACOS)". Headers are normalized (lowercased, currency symbols and
bracketed suffixes stripped) and matched against known aliases, so reports
from any marketplace parse without configuration.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

from .models import SearchTermRow

# normalized header fragment -> SearchTermRow field
_HEADER_PATTERNS = [
    (r"^campaign( name)?$", "campaign"),
    (r"^ad group( name)?$", "ad_group"),
    (r"^targeting$", "targeting"),
    (r"^match type$", "match_type"),
    (r"^(customer )?search term$", "search_term"),
    (r"^impressions$", "impressions"),
    (r"^clicks$", "clicks"),
    (r"^(spend|cost)$", "spend"),
    (r"total sales$", "sales"),
    (r"total orders", "orders"),
    (r"total units", "units"),
]

_NUMERIC_FIELDS = {"impressions", "clicks", "orders", "units"}
_MONEY_FIELDS = {"spend", "sales"}

_STRIP_RE = re.compile(r"[₹$€£,%]|\((?:[^)]*)\)")


def normalize_header(header: str) -> str:
    """Lowercase and strip currency symbols and bracketed suffixes."""
    cleaned = _STRIP_RE.sub("", header.lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def map_headers(headers: list[str]) -> dict[int, str]:
    """Map column indexes to SearchTermRow field names."""
    mapping: dict[int, str] = {}
    for idx, header in enumerate(headers):
        normalized = normalize_header(str(header or ""))
        for pattern, fieldname in _HEADER_PATTERNS:
            if re.search(pattern, normalized) and fieldname not in mapping.values():
                mapping[idx] = fieldname
                break
    return mapping


def _parse_number(value, as_int: bool = False):
    if value is None:
        return 0 if as_int else 0.0
    if isinstance(value, (int, float)):
        return int(value) if as_int else float(value)
    text = _STRIP_RE.sub("", str(value)).strip()
    if not text or text == "-":
        return 0 if as_int else 0.0
    try:
        return int(float(text)) if as_int else float(text)
    except ValueError:
        return 0 if as_int else 0.0


def _rows_from_table(table) -> list[SearchTermRow]:
    table = iter(table)
    try:
        headers = list(next(table))
    except StopIteration:
        return []
    mapping = map_headers(headers)
    required = {"search_term", "clicks", "spend"}
    missing = required - set(mapping.values())
    if missing:
        raise ValueError(
            f"Report is missing required columns: {sorted(missing)}. "
            "Expected an Amazon Ads search term report export."
        )

    rows = []
    for raw in table:
        row = SearchTermRow()
        for idx, fieldname in mapping.items():
            value = raw[idx] if idx < len(raw) else None
            if fieldname in _NUMERIC_FIELDS:
                setattr(row, fieldname, _parse_number(value, as_int=True))
            elif fieldname in _MONEY_FIELDS:
                setattr(row, fieldname, _parse_number(value))
            else:
                setattr(row, fieldname, str(value).strip() if value is not None else "")
        if row.search_term or row.impressions or row.clicks:
            rows.append(row)
    return rows


def read_report(path: str | Path) -> list[SearchTermRow]:
    """Read a search term report (.csv, .tsv, or .xlsx) into rows."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".csv", ".tsv"):
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return _rows_from_table(csv.reader(fh, delimiter=delimiter))
    if suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            return _rows_from_table(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
    raise ValueError(f"Unsupported report format '{suffix}'. Use .csv, .tsv, or .xlsx.")
