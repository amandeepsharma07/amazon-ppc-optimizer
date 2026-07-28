"""Write recommendations to an Excel workbook or CSV files."""

from __future__ import annotations

import csv
import dataclasses
from pathlib import Path

_SHEETS = [
    ("Campaign Summary", "summary"),
    ("Bid Recommendations", "bids"),
    ("Negative Keywords", "negatives"),
    ("Keyword Harvest", "harvest"),
]

_PERCENT_FIELDS = {"acos", "ctr", "cvr"}


def _headers_and_rows(records: list) -> tuple[list[str], list[list]]:
    if not records:
        return [], []
    fields = [f.name for f in dataclasses.fields(records[0])]
    headers = [f.replace("_", " ").title() for f in fields]
    rows = [[getattr(r, f) for f in fields] for r in records]
    return headers, rows


def write_xlsx(path: str | Path, summary, bids, negatives, harvest) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = {"summary": summary, "bids": bids, "negatives": negatives, "harvest": harvest}
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="232F3E")  # Amazon navy

    for title, key in _SHEETS:
        sheet = workbook.create_sheet(title)
        records = data[key]
        headers, rows = _headers_and_rows(records)
        if not headers:
            sheet["A1"] = "No recommendations in this category."
            continue
        fields = [f.name for f in dataclasses.fields(records[0])]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            sheet.append(row)
        for col_idx, fieldname in enumerate(fields, start=1):
            letter = get_column_letter(col_idx)
            width = max(len(headers[col_idx - 1]) + 2, 12)
            for row in rows[:50]:
                width = max(width, min(len(str(row[col_idx - 1])) + 2, 48))
            sheet.column_dimensions[letter].width = width
            if fieldname in _PERCENT_FIELDS:
                for cell in sheet[letter][1:]:
                    cell.number_format = "0.0%"
        sheet.freeze_panes = "A2"

    path = Path(path)
    workbook.save(path)
    return path


def write_csvs(directory: str | Path, summary, bids, negatives, harvest) -> list[Path]:
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)
    data = {"summary": summary, "bids": bids, "negatives": negatives, "harvest": harvest}
    written = []
    for title, key in _SHEETS:
        headers, rows = _headers_and_rows(data[key])
        path = directory / f"{key}.csv"
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            if headers:
                writer.writerow(headers)
                writer.writerows(rows)
        written.append(path)
    return written
