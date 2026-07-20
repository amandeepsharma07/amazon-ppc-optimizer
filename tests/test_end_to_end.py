from pathlib import Path

from ppcopt.cli import main
from ppcopt.sample import generate_sample


def test_sample_then_optimize_xlsx(tmp_path: Path, capsys):
    report = tmp_path / "sample.csv"
    generate_sample(report, rows=200)
    output = tmp_path / "recs.xlsx"

    exit_code = main(["optimize", str(report), "--target-acos", "30",
                      "-o", str(output)])
    assert exit_code == 0
    assert output.exists()

    from openpyxl import load_workbook

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Campaign Summary", "Bid Recommendations",
        "Negative Keywords", "Keyword Harvest",
    ]
    summary = workbook["Campaign Summary"]
    assert summary.max_row > 1  # has data beyond the header

    captured = capsys.readouterr()
    assert "Bid recommendations" in captured.out


def test_optimize_csv_output(tmp_path: Path):
    report = tmp_path / "sample.csv"
    generate_sample(report, rows=100)
    out_dir = tmp_path / "out"

    exit_code = main(["optimize", str(report), "--format", "csv",
                      "-o", str(out_dir)])
    assert exit_code == 0
    names = sorted(p.name for p in out_dir.glob("*.csv"))
    assert names == ["bids.csv", "harvest.csv", "negatives.csv", "summary.csv"]


def test_target_acos_accepts_percent_and_fraction(tmp_path: Path):
    from ppcopt.cli import _parse_acos

    assert _parse_acos("30") == 0.30
    assert _parse_acos("30%") == 0.30
    assert _parse_acos("0.30") == 0.30
